from decimal import Decimal
from uuid import UUID
from fastapi import APIRouter, Depends, Query
from sqlalchemy.orm import Session
from sqlalchemy import func, case, desc, extract
from typing import List, Optional
from datetime import datetime, timedelta
from app.database import get_db_session
from app.models.base import (
    Product, ProductCategory, StorageLocation, 
    Ticket, SaleItem, SaleState, Payment, PaymentMethod,
    Purchase, PurchaseItem, Supplier, PurchaseState,
    CashSession, InventoryMovement, MovementType, ProductType
)

router = APIRouter()


def parse_date(value: Optional[str]) -> Optional[datetime]:
    """Parsea fechas en formato 'YYYY-MM-DD' o ISO completo desde el frontend."""
    if not value:
        return None
    try:
        # Intenta primero ISO completo
        return datetime.fromisoformat(value)
    except ValueError:
        pass
    try:
        # Intenta solo fecha YYYY-MM-DD
        return datetime.strptime(value, "%Y-%m-%d")
    except ValueError:
        return None


# ── SALES REPORTS ─────────────────────────────────────────────────────────────

@router.get("/sales/summary")
def get_sales_summary(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db_session)
):
    start_dt = parse_date(start_date) or datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
    end_dt = parse_date(end_date) or datetime.utcnow()
    start_date = start_dt
    end_date = end_dt

    # Base query: incluir VALIDATED y PAID (ambos estados son ventas realizadas)
    VALID_STATES = [SaleState.VALIDATED, SaleState.PAID]

    query = db.query(Ticket).filter(
        Ticket.date_created >= start_date,
        Ticket.date_created <= end_date,
        Ticket.state.in_(VALID_STATES),
        Ticket.is_refunded == False  # Excluir ventas reembolsadas
    )

    # 1. Gross Sales
    gross_sales = query.with_entities(func.sum(Ticket.total_amount)).scalar() or Decimal('0.00')

    # 2. Avg Ticket
    total_tickets = query.count()
    avg_ticket = gross_sales / total_tickets if total_tickets > 0 else Decimal('0.00')

    # 3. Digital Payments (Card + Transfer)
    digital_sales = db.query(func.sum(Payment.amount)).join(Ticket).filter(
        Ticket.date_created >= start_date,
        Ticket.date_created <= end_date,
        Ticket.state.in_(VALID_STATES),
        Ticket.is_refunded == False,
        Payment.payment_method.in_([PaymentMethod.CARD, PaymentMethod.TRANSFER])
    ).scalar() or Decimal('0.00')

    # 4a. Sales by Hour (para vista de día único)
    hourly_data_raw = query.with_entities(Ticket.date_created, Ticket.total_amount).all()
    
    sales_by_hour = {}
    for t in hourly_data_raw:
        hour = t.date_created.hour
        hour_label = f"{hour:02d}:00"
        sales_by_hour[hour_label] = sales_by_hour.get(hour_label, Decimal('0')) + (t.total_amount or Decimal('0'))
    
    # Rango 8 a 22
    chart_data = []
    for h in range(8, 23):
        label = f"{h:02d}:00"
        chart_data.append({"hour": label, "ventas": float(sales_by_hour.get(label, 0))})

    # 4b. Sales by Day + Ganancia bruta por día (ventas - costo)
    # Obtenemos items de venta en el rango para calcular costo
    sale_items_in_range = (
        db.query(SaleItem)
        .join(Ticket)
        .filter(
            Ticket.date_created >= start_date,
            Ticket.date_created <= end_date,
            Ticket.state.in_(VALID_STATES),
            Ticket.is_refunded == False
        )
        .all()
    )

    daily_sales: dict = {}      # fecha -> ventas totales
    daily_cost: dict = {}       # fecha -> costo total

    for item in sale_items_in_range:
        if not item.ticket:
            continue
        day_key = item.ticket.date_created.strftime("%Y-%m-%d")
        daily_sales[day_key] = daily_sales.get(day_key, Decimal('0')) + (item.subtotal or Decimal('0'))
        product_cost = (item.product.cost if item.product else Decimal('0')) or Decimal('0')
        daily_cost[day_key] = daily_cost.get(day_key, Decimal('0')) + product_cost * item.quantity

    # Generar rango completo de días
    from datetime import date as date_type
    delta = (end_dt.date() - start_dt.date()).days
    daily_chart_data = []
    for i in range(delta + 1):
        d = (start_dt + timedelta(days=i)).strftime("%Y-%m-%d")
        ventas = float(daily_sales.get(d, 0))
        costo = float(daily_cost.get(d, 0))
        ganancia = ventas - costo
        # Etiqueta amigable: dd/mm
        label = f"{d[8:10]}/{d[5:7]}"
        daily_chart_data.append({
            "day": label,
            "date": d,
            "ventas": ventas,
            "ganancia": ganancia,
        })

    # 5. Recent Transactions — ordenar por date_created (campo correcto del modelo)
    transactions = query.order_by(desc(Ticket.date_created)).limit(10).all()
    trans_list = []
    for t in transactions:
        # Determinar método de pago dominante
        method = "Mixto"
        if len(t.payments) == 1:
            method = t.payments[0].payment_method.value
        elif len(t.payments) > 1:
            method = "mixto"
        
        # Cajero desde la sesión
        cajero = "S/A"
        if t.session and t.session.user_id:
            cajero = t.session.user_id
        
        trans_list.append({
            "id": t.ticket_number,
            "cajero": cajero,
            "metodo": method,
            "total": float(t.total_amount),
            "estado": t.state.value,
            "date": t.date_created.isoformat(),
            "document_type": t.document_type,
            "comment": t.comment
        })

    return {
        "kpis": {
            "gross_sales": float(gross_sales),
            "avg_ticket": float(avg_ticket),
            "digital_sales": float(digital_sales),
            "total_tickets": total_tickets
        },
        "chart_data": chart_data,
        "daily_chart_data": daily_chart_data,
        "recent_transactions": trans_list
    }

@router.get("/sales/profitability")
def get_profitability(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db_session)
):
    start_dt = parse_date(start_date) or datetime.utcnow().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    end_dt = parse_date(end_date) or datetime.utcnow()
    start_date = start_dt
    end_date = end_dt

    VALID_STATES = [SaleState.VALIDATED, SaleState.PAID]

    items = db.query(SaleItem).join(Ticket).filter(
        Ticket.date_created >= start_date,
        Ticket.date_created <= end_date,
        Ticket.state.in_(VALID_STATES),
        Ticket.is_refunded == False
    ).all()

    total_sales = Decimal('0')
    total_cost = Decimal('0')
    
    product_stats = {}  # product_id -> {sales, cost, qty, name, category}

    for item in items:
        product = item.product
        if not product:
            continue

        cost = (product.cost or Decimal('0')) * item.quantity
        sale = item.subtotal or Decimal('0')
        
        total_sales += sale
        total_cost += cost
        
        if product.id not in product_stats:
            product_stats[product.id] = {
                "name": product.name,
                "category": product.category_rel.name if product.category_rel else "Sin Categoría",
                "sales": Decimal('0'),
                "cost": Decimal('0'),
                "qty": Decimal('0'),
            }
        
        product_stats[product.id]["sales"] += sale
        product_stats[product.id]["cost"] += cost
        product_stats[product.id]["qty"] += item.quantity

    total_margin = total_sales - total_cost
    global_margin_pct = float((total_margin / total_sales * 100) if total_sales > 0 else 0)

    # Best Product por margen absoluto
    best_product = None
    max_margin = Decimal('-999999')

    detailed_products = []
    for pid, stats in product_stats.items():
        margin = stats["sales"] - stats["cost"]
        if margin > max_margin:
            max_margin = margin
            best_product = stats
        
        avg_price = float(stats["sales"] / stats["qty"]) if stats["qty"] > 0 else 0
        avg_cost = float(stats["cost"] / stats["qty"]) if stats["qty"] > 0 else 0
        
        detailed_products.append({
            "descripcion": stats["name"],
            "categoria": stats["category"],
            "costo": avg_cost,
            "venta": avg_price,
            "margen": float(margin),
            "unidades": float(stats["qty"])
        })

    # Ordenar por margen descendente
    detailed_products.sort(key=lambda x: x["margen"], reverse=True)

    # Distribución por categoría (sin doble suma)
    cat_stats = {}
    for pid, stats in product_stats.items():
        cat = stats["category"]
        if cat not in cat_stats:
            cat_stats[cat] = {"sales": Decimal('0'), "cost": Decimal('0')}
        cat_stats[cat]["sales"] += stats["sales"]
        cat_stats[cat]["cost"] += stats["cost"]

    chart_data = []
    for cat, data in cat_stats.items():
        chart_data.append({
            "categoria": cat,
            "costo": float(data["cost"]),
            "venta": float(data["sales"])
        })

    best_product_name = best_product["name"] if best_product else "N/A"
    best_product_margin_pct = 0.0
    if best_product and best_product["sales"] > 0:
        best_margin = best_product["sales"] - best_product["cost"]
        best_product_margin_pct = float(best_margin / best_product["sales"] * 100)

    return {
        "kpis": {
            "total_margin": float(total_margin),
            "global_margin_pct": global_margin_pct,
            "best_product": best_product_name,
            "best_product_margin": best_product_margin_pct
        },
        "chart_data": chart_data,
        "table_data": detailed_products
    }

@router.get("/sales/cash_reports")
def get_cash_reports(db: Session = Depends(get_db_session)):
    # Usar los campos REALES del modelo CashSession:
    # opened_at, closed_at, opening_balance, closing_balance, status, user_id, cash_register
    sessions_db = db.query(CashSession).order_by(desc(CashSession.opened_at)).limit(10).all()
    
    data = []
    total_diff = Decimal('0')
    cash_in_hand = Decimal('0')
    
    for s in sessions_db:
        closing = s.closing_balance or Decimal('0')
        opening = s.opening_balance or Decimal('0')
        
        # difference calculado como: closing - expected
        # expected_balance está guardado en el modelo
        expected = s.expected_balance or opening
        diff = s.difference or (closing - expected)
        
        total_diff += diff
        
        # Si la sesión está abierta, sumar el opening como efectivo en curso
        if s.status == "open":
            cash_in_hand += opening
        else:
            cash_in_hand += closing
        
        # Nombre de la caja desde la relación cash_register
        caja_name = s.cash_register.name if s.cash_register else "Caja"
        
        data.append({
            "usuario": s.user_id or "Anónimo",
            "caja": caja_name,
            "aperturaHora": s.opened_at.strftime("%H:%M") if s.opened_at else "--",
            "aperturaMonto": float(opening),
            "cierreHora": s.closed_at.strftime("%H:%M") if s.closed_at else "En curso",
            "cierreMonto": float(closing),
            "diferencia": float(diff),
            "estado": s.status
        })
        
    return {
        "kpis": {
            "cash_in_hand": float(cash_in_hand),
            "total_difference": float(total_diff)
        },
        "sessions": data
    }


# ── INVENTORY REPORTS ─────────────────────────────────────────────────────────

@router.get("/inventory/summary")
def get_inventory_summary(
    aisle: Optional[str] = None,
    category: Optional[str] = None,
    db: Session = Depends(get_db_session)
):
    # Query Products (Excluyendo SERVICIOS)
    q = db.query(Product).filter(
        Product.is_active == True,
        Product.product_type != ProductType.SERVICE
    )
    
    if category and category != "all":
        q = q.join(ProductCategory).filter(ProductCategory.name == category)
    
    if aisle and aisle != "all":
        q = q.join(StorageLocation).filter(StorageLocation.path.contains(aisle))

    products = q.all()
    
    total_valuation = Decimal('0')
    low_stock_count = 0
    occupied_locs = set()
    
    audit_list = []
    cat_distribution = {}
    
    for p in products:
        val = (p.stock_quantity or Decimal('0')) * (p.cost or Decimal('0'))
        total_valuation += val
        
        if (p.stock_quantity or 0) <= (p.min_stock or 0):
            low_stock_count += 1
            
        if p.location:
            occupied_locs.add(p.location.id)
            coord = p.location.name
        else:
            coord = "Sin Ubicación"
            
        cat_name = p.category_rel.name if p.category_rel else "Sin Categoría"
        cat_color = p.category_rel.color if p.category_rel else "#ccc"
        
        cat_distribution.setdefault(cat_name, {"val": Decimal('0'), "color": cat_color})
        cat_distribution[cat_name]["val"] += (p.stock_quantity or Decimal('0'))
        
        audit_list.append({
            "name": p.name,
            "category": cat_name,
            "coord": coord,
            "stock": float(p.stock_quantity or 0),
            "minStock": float(p.min_stock or 0),
            "aisle": "N/A"
        })
        
    donut_data = [
        {"name": k, "value": float(v["val"]), "fill": v["color"]} 
        for k, v in cat_distribution.items()
    ]
    
    return {
        "kpis": {
            "total_valuation": float(total_valuation),
            "low_stock": low_stock_count,
            "occupied_locations": len(occupied_locs)
        },
        "donut_data": donut_data,
        "audit_table": audit_list
    }

# ── PURCHASES REPORTS ─────────────────────────────────────────────────────────

@router.get("/purchases/summary")
def get_purchases_summary(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db_session)
):
    start_dt = parse_date(start_date) or datetime.utcnow().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    end_dt = parse_date(end_date) or datetime.utcnow()
    start_date = start_dt
    end_date = end_dt
    
    purchases_q = db.query(Purchase).filter(
        Purchase.state == PurchaseState.CONFIRMED,
        Purchase.date_created >= start_date,
        Purchase.date_created <= end_date
    )
    
    total_invested = sum(p.total_cost or Decimal('0') for p in purchases_q.all())
    
    # Pendientes
    pending_count = db.query(Purchase).filter(Purchase.state == PurchaseState.DRAFT).count()
    
    # Top Supplier
    suppliers_vol = {}
    all_confirmed = db.query(Purchase).filter(Purchase.state == PurchaseState.CONFIRMED).all()
    for p in all_confirmed:
        s_name = p.supplier.name if p.supplier else "Desconocido"
        suppliers_vol[s_name] = suppliers_vol.get(s_name, Decimal('0')) + (p.total_cost or Decimal('0'))
        
    sorted_suppliers = sorted(suppliers_vol.items(), key=lambda x: x[1], reverse=True)
    top_supplier = {"name": "N/A", "volume": 0}
    if sorted_suppliers:
        top_supplier = {"name": sorted_suppliers[0][0], "volume": float(sorted_suppliers[0][1])}
        
    # Chart Data (Top 5)
    chart_data = [{"name": s[0], "volume": float(s[1])} for s in sorted_suppliers[:5]]
    
    # Movimientos recientes
    recent_items = db.query(PurchaseItem).join(Purchase).order_by(desc(Purchase.date_created)).limit(20).all()
    movements = []
    for item in recent_items:
        movements.append({
            "fecha": item.purchase.date_created.strftime("%Y-%m-%d"),
            "proveedor": item.purchase.supplier.name if item.purchase.supplier else "?",
            "producto": item.product.name if item.product else "?",
            "cantidad": float(item.quantity),
            "costoUnit": float(item.unit_cost)
        })
        
    return {
        "kpis": {
            "total_invested": float(total_invested),
            "pending_orders": pending_count,
            "top_supplier": top_supplier
        },
        "chart_data": chart_data,
        "movements": movements
    }


# ── SALES EXPORT ───────────────────────────────────────────────────────────────

@router.get("/sales/export")
def export_sales_excel(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db_session)
):
    """
    Genera un Excel de ventas con 3 hojas:
      1. Resumen  – KPIs + desglose por método de pago
      2. Transacciones – cada ticket con cajero, método, total, estado
      3. Ítems Vendidos – producto, cantidad, precio unitario, costo, margen
    """
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from fastapi.responses import StreamingResponse

    start_dt = parse_date(start_date) or datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
    end_dt   = parse_date(end_date)   or datetime.utcnow()

    VALID_STATES = [SaleState.VALIDATED, SaleState.PAID]

    tickets = (
        db.query(Ticket)
        .filter(
            Ticket.date_created >= start_dt,
            Ticket.date_created <= end_dt,
            Ticket.state.in_(VALID_STATES),
            Ticket.is_refunded == False,
        )
        .order_by(desc(Ticket.date_created))
        .all()
    )

    # ── Helpers de estilo ────────────────────────────────────────────────────
    HEADER_FILL   = PatternFill("solid", fgColor="1A1A2E")   # Azul oscuro
    HEADER_FONT   = Font(bold=True, color="FFFFFF", size=11)
    ACCENT_FILL   = PatternFill("solid", fgColor="16213E")
    SUBHEAD_FILL  = PatternFill("solid", fgColor="0F3460")
    SUBHEAD_FONT  = Font(bold=True, color="FFFFFFDD", size=10)
    TOTAL_FILL    = PatternFill("solid", fgColor="E2F0D9")
    TOTAL_FONT    = Font(bold=True, size=11)
    thin_side     = Side(style="thin", color="CCCCCC")
    BORDER        = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    def style_header(ws, row, cols):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill   = HEADER_FILL
            cell.font   = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = BORDER

    def style_row(ws, row, cols, alt=False):
        fill = PatternFill("solid", fgColor="F7F7F7") if alt else PatternFill("solid", fgColor="FFFFFF")
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill   = fill
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center")

    def auto_width(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value or "")))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 50)

    # ── Workbook ─────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()

    label = f"{start_dt.strftime('%d/%m/%Y')} – {end_dt.strftime('%d/%m/%Y')}"

    # ═══════════════════════════════════════════════════════════════
    # HOJA 1 — RESUMEN
    # ═══════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Resumen"

    # Título
    ws1.merge_cells("A1:D1")
    ws1["A1"] = f"📊 Reporte de Ventas   {label}"
    ws1["A1"].font   = Font(bold=True, size=14, color="1A1A2E")
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 32
    ws1.merge_cells("A2:D2")
    ws1["A2"] = f"Generado: {datetime.utcnow().strftime('%d/%m/%Y %H:%M')} UTC"
    ws1["A2"].font   = Font(italic=True, size=9, color="888888")
    ws1["A2"].alignment = Alignment(horizontal="center")
    ws1.row_dimensions[2].height = 16

    # KPIs
    total_ventas  = sum(t.total_amount or Decimal('0') for t in tickets)
    total_tickets = len(tickets)
    avg_ticket    = total_ventas / total_tickets if total_tickets else Decimal('0')

    pay_totals: dict = {}
    for t in tickets:
        for p in t.payments:
            m = p.payment_method.value
            pay_totals[m] = pay_totals.get(m, Decimal('0')) + (p.amount or Decimal('0'))

    ws1.row_dimensions[3].height = 24
    ws1["A4"] = "KPI"; ws1["B4"] = "Valor"
    style_header(ws1, 4, 2)
    ws1.row_dimensions[4].height = 22

    kpis = [
        ("Venta Bruta Total",   f"${float(total_ventas):,.2f}"),
        ("Nº de Transacciones", total_tickets),
        ("Ticket Promedio",     f"${float(avg_ticket):,.2f}"),
    ]
    for i, (k, v) in enumerate(kpis, start=5):
        ws1[f"A{i}"] = k
        ws1[f"B{i}"] = v
        style_row(ws1, i, 2, alt=(i % 2 == 0))

    # Métodos de pago
    row = 5 + len(kpis) + 1
    ws1.merge_cells(f"A{row}:D{row}")
    ws1[f"A{row}"] = "Métodos de Pago"
    ws1[f"A{row}"].fill = SUBHEAD_FILL
    ws1[f"A{row}"].font = SUBHEAD_FONT
    ws1[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[row].height = 22
    row += 1

    ws1[f"A{row}"] = "Método"; ws1[f"B{row}"] = "Total"
    style_header(ws1, row, 2)
    ws1.row_dimensions[row].height = 20
    row += 1

    for j, (method, amount) in enumerate(pay_totals.items()):
        ws1[f"A{row}"] = method.capitalize()
        ws1[f"B{row}"] = f"${float(amount):,.2f}"
        style_row(ws1, row, 2, alt=(j % 2 == 0))
        row += 1

    # Fila total
    ws1[f"A{row}"] = "TOTAL"
    ws1[f"B{row}"] = f"${float(sum(pay_totals.values())):,.2f}"
    for c in ["A", "B"]:
        ws1[f"{c}{row}"].fill = TOTAL_FILL
        ws1[f"{c}{row}"].font = TOTAL_FONT
        ws1[f"{c}{row}"].border = BORDER
        ws1[f"{c}{row}"].alignment = Alignment(horizontal="center")

    auto_width(ws1)
    for col in ["A", "B"]:
        ws1.column_dimensions[col].width = max(ws1.column_dimensions[col].width, 22)

    # ═══════════════════════════════════════════════════════════════
    # HOJA 2 — TRANSACCIONES
    # ═══════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Transacciones")

    headers2 = ["ID Venta", "Fecha", "Hora", "Cajero", "Método de Pago", "Comprobante", "Total", "Estado", "Comentario"]
    for j, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=j, value=h)
    style_header(ws2, 1, len(headers2))
    ws2.row_dimensions[1].height = 22

    for i, t in enumerate(tickets, start=2):
        method = "Mixto"
        if len(t.payments) == 1:
            method = t.payments[0].payment_method.value.capitalize()
        elif len(t.payments) > 1:
            method = "Mixto"

        cajero = "S/A"
        if t.session and t.session.user_id:
            cajero = t.session.user_id

        row_data = [
            t.ticket_number,
            t.date_created.strftime("%d/%m/%Y"),
            t.date_created.strftime("%H:%M"),
            cajero,
            method,
            (t.document_type or "boleta").capitalize(),
            float(t.total_amount or 0),
            t.state.value.capitalize(),
            t.comment or ""
        ]
        for j, val in enumerate(row_data, 1):
            ws2.cell(row=i, column=j, value=val)

        # Color fondo total (columna 7 ahora)
        ws2.cell(row=i, column=7).number_format = '"$"#,##0.00'
        style_row(ws2, i, len(headers2), alt=(i % 2 == 0))

    auto_width(ws2)

    # ═══════════════════════════════════════════════════════════════
    # HOJA 3 — ÍTEMS VENDIDOS
    # ═══════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Ítems Vendidos")

    headers3 = ["ID Venta", "Fecha", "Producto", "Categoría", "Cantidad", "Precio Unit.", "Subtotal", "Costo Unit.", "Costo Total", "Margen ($)", "Margen (%)"]
    for j, h in enumerate(headers3, 1):
        ws3.cell(row=1, column=j, value=h)
    style_header(ws3, 1, len(headers3))
    ws3.row_dimensions[1].height = 22

    row3 = 2
    for t in tickets:
        for item in t.items:
            product   = item.product
            cost_unit = (product.cost if product else Decimal('0')) or Decimal('0')
            qty       = item.quantity or Decimal('0')
            subtotal  = item.subtotal or Decimal('0')
            costo_tot = cost_unit * qty
            margen    = subtotal - costo_tot
            margen_pct = float(margen / subtotal * 100) if subtotal > 0 else 0

            cat = product.category_rel.name if (product and product.category_rel) else "Sin Categoría"
            row_data = [
                t.ticket_number,
                t.date_created.strftime("%d/%m/%Y"),
                product.name if product else "?",
                cat,
                float(qty),
                float(item.unit_price or 0),
                float(subtotal),
                float(cost_unit),
                float(costo_tot),
                float(margen),
                round(margen_pct, 1),
            ]
            for j, val in enumerate(row_data, 1):
                ws3.cell(row=row3, column=j, value=val)

            # Formato moneda
            for col_idx in [6, 7, 8, 9, 10]:
                ws3.cell(row=row3, column=col_idx).number_format = '"$"#,##0.00'
            ws3.cell(row=row3, column=11).number_format = '0.0"%"'

            style_row(ws3, row3, len(headers3), alt=(row3 % 2 == 0))

            # Verde/Rojo según margen
            margen_cell = ws3.cell(row=row3, column=10)
            if float(margen) >= 0:
                margen_cell.font = Font(color="1A8055", bold=True)
            else:
                margen_cell.font = Font(color="C0392B", bold=True)

            row3 += 1

    auto_width(ws3)

    # ── Serializar y retornar ─────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"ventas_{start_dt.strftime('%Y%m%d')}_{end_dt.strftime('%Y%m%d')}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ── INVENTORY EXPORT ────────────────────────────────────────────────────────────

@router.get("/inventory/export")
def export_inventory_excel(
    category: Optional[str] = None,
    aisle: Optional[str] = None,
    db: Session = Depends(get_db_session)
):
    """
    Genera un Excel del estado actual del inventario.
    Incluye: Categoría, Producto, Stock, Costo, Precio, Valorización, Margen y Ubicación.
    """
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from fastapi.responses import StreamingResponse

    q = db.query(Product).join(ProductCategory, isouter=True).join(StorageLocation, isouter=True)
    if category and category != "all":
        q = q.filter(ProductCategory.name == category)
    if aisle:
        q = q.filter(StorageLocation.name.ilike(f"%{aisle}%"))
    
    products = q.order_by(ProductCategory.name, Product.name).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario"

    # Estilos (reutilizando lógica premium)
    HEADER_FILL = PatternFill("solid", fgColor="1A1A2E")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    thin_side   = Side(style="thin", color="CCCCCC")
    BORDER      = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # Título
    ws.merge_cells("A1:I1")
    ws["A1"] = "📦 Reporte de Estado de Inventario"
    ws["A1"].font = Font(bold=True, size=14, color="1A1A2E")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:I2")
    ws["A2"] = f"Generado: {datetime.utcnow().strftime('%d/%m/%Y %H:%M')} | Filtros: {category or 'General'} / {aisle or 'Todas las ubic.'}"
    ws["A2"].font = Font(italic=True, size=9, color="888888")
    ws["A2"].alignment = Alignment(horizontal="center")

    headers = ["Categoría", "Producto", "Stock Actual", "Mínimo", "Costo Unit.", "Precio Venta", "Valor Total (Costo)", "Margen Esperado (%)", "Ubicación"]
    for j, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=j, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER
    ws.row_dimensions[4].height = 22

    total_val = 0
    row_idx = 5
    for p in products:
        cost = float(p.cost or 0)
        price = float(p.price or 0)
        stock = float(p.stock_quantity or 0)
        val_sub = cost * stock
        total_val += val_sub
        
        margen = ((price - cost) / price * 100) if price > 0 else 0
        
        row_data = [
            p.category_rel.name if p.category_rel else "Sin Cat.",
            p.name,
            stock,
            float(p.min_stock or 0),
            cost,
            price,
            val_sub,
            round(margen, 1),
            p.location.name if p.location else "S/U"
        ]
        
        for j, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=j, value=val)
            cell.border = BORDER
            if j in [3, 4, 5, 6, 7]: cell.alignment = Alignment(horizontal="right")
            if j in [5, 6, 7]: cell.number_format = '"$"#,##0.00'
            if j == 8: cell.number_format = '0.0"%"'
            
        # Color rojo si stock <= min
        if stock <= float(p.min_stock or 0):
            ws.cell(row=row_idx, column=3).font = Font(color="C0392B", bold=True)

        row_idx += 1

    # Fila Resumen Final
    ws.cell(row=row_idx, column=6, value="VALOR TOTAL:").font = Font(bold=True)
    val_cell = ws.cell(row=row_idx, column=7, value=total_val)
    val_cell.font = Font(bold=True)
    val_cell.number_format = '"$"#,##0.00'
    val_cell.fill = PatternFill("solid", fgColor="E2F0D9")

    # Auto-ajuste de columnas
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try: max_len = max(max_len, len(str(cell.value or "")))
            except: pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"inventario_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ── PURCHASES EXPORT ────────────────────────────────────────────────────────────

@router.get("/purchases/export")
def export_purchases_excel(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db_session)
):
    """
    Genera un Excel de compras con 2 hojas:
      1. Resumen por Proveedor
      2. Detalle de Ítems Comprados
    """
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from fastapi.responses import StreamingResponse

    start_dt = parse_date(start_date) or datetime.utcnow().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    end_dt   = parse_date(end_date)   or datetime.utcnow()

    purchases = (
        db.query(Purchase)
        .filter(
            Purchase.date_created >= start_dt,
            Purchase.date_created <= end_dt,
            Purchase.state == PurchaseState.CONFIRMED
        )
        .order_by(desc(Purchase.date_created))
        .all()
    )

    wb = openpyxl.Workbook()
    
    # Estilos comunes
    HEADER_FILL = PatternFill("solid", fgColor="1A1A2E")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    thin_side   = Side(style="thin", color="CCCCCC")
    BORDER      = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    def style_header(ws, cols):
        for c in range(1, cols + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
            cell.border = BORDER

    def auto_width(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try: max_len = max(max_len, len(str(cell.value or "")))
                except: pass
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 40)

    # ═══════════════════════════════════════════════════════════════
    # HOJA 1 — RESUMEN PROVEEDORES
    # ═══════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Resumen Proveedores"
    
    headers1 = ["Proveedor", "Nº Compras", "Inversión Total"]
    for j, h in enumerate(headers1, 1):
        ws1.cell(row=1, column=j, value=h)
    style_header(ws1, len(headers1))

    prov_agg = {}
    for p in purchases:
        name = p.supplier.name if p.supplier else "S/P"
        data = prov_agg.get(name, {"count": 0, "total": Decimal('0')})
        data["count"] += 1
        data["total"] += (p.total_cost or Decimal('0'))
        prov_agg[name] = data

    for i, (name, stats) in enumerate(prov_agg.items(), 2):
        ws1.cell(row=i, column=1, value=name)
        ws1.cell(row=i, column=2, value=stats["count"])
        cell_val = ws1.cell(row=i, column=3, value=float(stats["total"]))
        cell_val.number_format = '"$"#,##0.00'
        
    auto_width(ws1)

    # ═══════════════════════════════════════════════════════════════
    # HOJA 2 — DETALLE DE ÍTEMS
    # ═══════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Detalle Items")
    headers2 = ["Fecha", "Proveedor", "Orden #", "Producto", "Cantidad", "Costo Unit.", "Subtotal"]
    for j, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=j, value=h)
    style_header(ws2, len(headers2))

    row_idx = 2
    for p in purchases:
        prov_name = p.supplier.name if p.supplier else "S/P"
        for item in p.items:
            qty = float(item.quantity or 0)
            cost = float(item.unit_cost or 0)
            ws2.cell(row=row_idx, column=1, value=p.date_created.strftime("%d/%m/%Y"))
            ws2.cell(row=row_idx, column=2, value=prov_name)
            ws2.cell(row=row_idx, column=3, value=p.id[:8].upper() if p.id else "N/A")
            ws2.cell(row=row_idx, column=4, value=item.product.name if item.product else "?")
            ws2.cell(row=row_idx, column=5, value=qty)
            c_u = ws2.cell(row=row_idx, column=6, value=cost)
            c_u.number_format = '"$"#,##0.00'
            s_t = ws2.cell(row=row_idx, column=7, value=qty * cost)
            s_t.number_format = '"$"#,##0.00'
            row_idx += 1

    auto_width(ws2)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"compras_{start_dt.strftime('%Y%m%d')}_{end_dt.strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
